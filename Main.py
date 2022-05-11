from calendar import month
from curses import meta
from gc import collect
from pickle import FALSE, NONE
from openpyxl import Workbook, workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Fill, NamedStyle
from openpyxl.styles.numbers import FORMAT_DATE_XLSX15, FORMAT_DATE_DDMMYY
from openpyxl.styles.colors import Color
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.utils import get_column_letter
from  collections import defaultdict
import Date, Autoxl
from Autoxl import res,headFont,TabColor,thin_border,alignCenter,tableFont
prod_ref_cells=[]
cost_ref_cells=[]




def main(orgReport, resReport, Day): #orgReport, resReport, Day
    #---------------Intial input of the two workbooks to work with-----------------#
    orgWB= orgReport #"original/OPR MCTR APR 11.xlsx"
    resWB= resReport#"results/Daily Eff. Report AutoTest.xlsx" #resReport
    #---------------Initializing Global Variables---------------------------#
    p1=Autoxl.autoxl()
    p1.original=load_workbook(orgWB)
    p1.result=load_workbook(resWB)
    Departments=p1.getDepart()
    Supervisors=p1.getSupers(len(Departments))
    Divisionrows=p1.getdeptDiv(len(Departments))
    #-------------------------Mode Select---------------------------------#
    print("\n\nPROGRAM START...\n\n")
    weekDay=Day #"L" #Day #input("What day of the week is it? Please input day as: M=Martes, W==Miercoles, J=Jueves or V=Viernes: ")
    #Modeselect='n'#input('\First entry of the week? Answer: y or n')

    if(weekDay=="L"):
        firstEntry=1
        date=p1.getDay()
        date=date.split("/")
        weekDate=Date.Date(int(date[1]),int(date[0]),2000+int(date[2]))
        endweekDate=weekDate.getWeek()

        #----------------Style for first entry of week-----------------------#
        #dateStyle=NamedStyle(name="dateformat",number_format=FORMAT_DATE_XLSX15)
        #p1.result.add_named_style(dateStyle)
        #--------------------------------------------------------------------#
        
        #Create Sheet and specifications
        sheetDate=" "+str(endweekDate.day)+"-"+res["Months"].get(endweekDate.month)
        sheetNames=[]
        #TODO Input the correct emplyees on the correct worksheets

        for i in range(0,len(Supervisors)):
            firstName=Supervisors[i].upper()
            sheetNames=sheetNames+[firstName+sheetDate]
            p1.result.create_sheet(sheetNames[i])
            ws=p1.result[sheetNames[i]]
            wsProp=ws.sheet_properties
            wsProp.tabColor=res["TabColors"].get(TabColor)

            #Sheet Title Creation
            createTitle(ws,weekDate,endweekDate,Departments,firstName)
            #Header Creation
            createHeader(ws,weekDate)

            print("\nStarting Employee Logging...")    
            iDs=p1.getIds(Divisionrows[i])
            employeeNames=p1.getNames(Divisionrows[i])
            effList=p1.getEff(Divisionrows[i]+1)

            #---------------Input new day entries---------------------------#
            inputIDname(ws,iDs,employeeNames,effList,Departments,i,firstEntry)
            #---------------------------------------------------------------# 
            #-------------------AutoFit Columns-----------------------------#
            fitTotext(firstEntry,ws,len(effList))  

            #Create Production, Overall Efficiency and Work Line Cost Table Reports
            #currRow=len(effList)
            #create_prod_rep(ws, currRow)


        #-----end for-------------------------------------------------------#  
        



        p1.result.save(resWB)
        print("\n\nPROGRAM END...")
    else:
        #Ask what day of the week is L=Lunes, M=Martes, W=Miercoles, J=Jueves, V=Viernes
        date=p1.getDay()
        date=date.split("/")
        weekDate=Date.Date(int(date[1]),int(date[0]),2000+int(date[2]))
        dayNum= weekDate.getdayNumber(weekDay)
        endweekDate=weekDate.getendWeek(weekDay)
        #-----------GET THE NAMES OF THE SHEETS TO EDIT----------------------------#
        sheetDate=" "+str(endweekDate.day)+"-"+res["Months"].get(endweekDate.month)
        sheetNames=[]
        for i in range(0,len(Supervisors)):
            two_diff=False
            firstName=Supervisors[i].upper()
            sheetNames=sheetNames+[firstName+sheetDate] #Gets the name of the sheet based on the supervisors in the original report
            ws=p1.result[sheetNames[i]]
            print("\nStarting Employee Logging...")    
            iDs=p1.getIds(Divisionrows[i])
            riDs=p1.getchildId(ws)
            employeeNames=p1.getNames(Divisionrows[i])
            effList=p1.getEff(Divisionrows[i]+1)
            DEBUG=0
            #----------ID comparison to see if new entries were made----------#
            iCount=0
            if(len(iDs)*2<len(riDs)):
                #---------Implementation of Dead Day, in case of very few employee Antendance--------------#
                p1.input_dead_day(ws,iDs,riDs,effList,dayNum)
            else:
                for (newElement,oldElement) in zip(iDs,riDs):
                    if(two_diff):
                        two_diff=False
                        continue
                        
                    #Check if a new entry has to be added by comparing the positions of the current 
                    # elements of each list
                    if(oldElement>newElement):
                        p1.setNewEntries(ws,employeeNames,iCount,iDs)
                        DEBUG+=1
                        riDs.insert(iCount,newElement)
                        iCount+=1
                    #Checks if a new entry is already in the Eff. Report, If so then this new entry 
                    # is added to the result Report
                    elif(oldElement<newElement):
                        try:
                            iDs.index(oldElement)
                        except:
                            effList.insert(iCount,0)
                            iDs.insert(iCount,oldElement)
                            employeeNames.insert(iCount," ")
                        try:
                            riDs.index(newElement)
                        except:
                            iCount+=1
                            p1.setNewEntries(ws,employeeNames,iCount,iDs)
                            DEBUG+=1
                            riDs.insert(iCount,newElement)
                            continue
                        #diff=abs(riDs.index(newElement)-iDs.index(newElement))
                        iCount+=1
                    else:
                        iCount+=1
                #---------------Check for Data Inconsistencies---------------------------#
                if(len(iDs)>len(riDs)):
                    diff=len(iDs)-len(riDs)
                    for i in range(0,diff):
                        p1.setNewEntries(ws,employeeNames,iCount,iDs)
                        DEBUG+=1
                        riDs.insert(iCount,iDs[iCount])
                        iCount+=1
                if(len(iDs)!=len(riDs)):
                    #Raise ValueError so that the workbook isn't saved with corrupted data
                    #raise ValueError('There were inconsistencies in the data to input with the data already in file')
                    print
                #---------------Input new day entries---------------------------#    
                inputIDname(ws,iDs,employeeNames,effList,Departments,i,dayNum)
                #---------------------------------------------------------------#
            #-------------------AutoFit Columns-----------------------------#
            fitTotext(dayNum, ws,len(effList))

            #--------------------Production, Overall Efficiency and Work Line Cost Table Reports----------------#
            if(weekDay=="V"):
                currRow=len(riDs)
                create_prod_rep(ws, currRow, i,sheetNames)
        #-----end for-------------------------------------------------------#
        p1.result.save(resWB)
        print("\n\nPROGRAM END...")    

def createTitle(ws,weekDate,endweekDate,Departments,supName):
    resValues=res["Title"]
    tMonths=res["Months"]
    print("Starting Title Creation...")
    for row in range(1,5):
        char1=get_column_letter(1)
        char2=get_column_letter(10)
        #print(char1,char2)
        ws.merge_cells(char1+str(row)+":"+char2+str(row))
        t=ws[char1+str(row)]
        if(row==3):
            t.value=resValues.get(row)+str(weekDate.day)+"-"+tMonths.get(weekDate.month)+"-"+str(weekDate.year)+" TO "+str(endweekDate.day)+"-"+tMonths.get(endweekDate.month)+"-"+str(endweekDate.year)
        elif(row==2):
            t.value=resValues.get(row)+"("+Departments[0]+")" #Departments[0] should be changed for when creating multiple sheets
        elif(row==4):
            t.value=resValues.get(row)+supName
        else:
            t.value=resValues.get(row)
        t.font=headFont
        t.alignment=alignCenter
    ws.pageSetUpPr = PageSetupProperties(fitToPage=True)
    print("Completed Title Creation...")

def createHeader(ws,weekDate):
    print("Starting Heading Creation...")
    #weekDate=Date(28,2,2022)
    resValues=res["Heading"]
    dayCount=weekDate.day
    for col in range(1,11):
        row=5
        if(col<4 or col>9):
            scol=get_column_letter(col)
            ws.merge_cells(scol+str(row)+":"+scol+str(row+2))
            t=ws[scol+str(row)]
            t.value=resValues.get(col)
            if(col>9):
                t.fill=PatternFill(start_color='F8B484', end_color='F8B484', fill_type="solid")
            else:
                t.fill=PatternFill(start_color='FFE49C', end_color='FFE49C', fill_type="solid")
            t.border=thin_border
        elif(col<9):
            for row in range(5,8):
                if(row==5):
                    #write day
                    dayCol=get_column_letter(col) 
                    t=ws[dayCol+str(row)]
                    t.value=resValues.get(col)
                elif(row==6):
                    #write date
                    t=ws[dayCol+str(row)]
                    if(col>4):
                        t.style="dateformat"
                        t.value="="+get_column_letter(col-1)+str(row)+"+1"
                    else:
                        t.style="dateformat"
                        t.value=str(weekDate.month)+"/"+str(dayCount)+"/"+str(weekDate.year)
                else:
                    t=ws[dayCol+str(row)]
                    t.value="Eff."
                cell_styling(t,"H")
                t.fill=PatternFill(start_color='FFE49C', end_color='FFE49C', fill_type="solid")
        t.font=headFont
        t.alignment=alignCenter
    print("Completed Heading Creation...")  
#Creates the following 3 tables that display Production, Line Efficiency and Cost per Unit

def create_prod_rep(ws, currRow, current_Supervisor,sheetNames):
    global prod_ref_cells, cost_ref_cells
    print("\nStarting Production Section Creation...")
    resValues=res["Prod_Eff"]
    resValuesH=res["Heading"]
    resTable=res["Prod_Cost_Title"]
    resCost=res["Cost"]
    org_row=currRow+7
    currRow+=10
    meta_cell=" "
    csr_Count=0

    for table in range(0,3):
        for col in range(1,11):
            if(col==1):
                col1=get_column_letter(col)
                col2=get_column_letter(col+1)
                ws.merge_cells(col1+str(currRow)+":"+col2+str(currRow+2))
                t=ws[col1+str(currRow)]
                t.value=resTable.get(table)
                cell_styling(t,"H")
            elif(col==3):
                if(table!=2):
                    col1=get_column_letter(col)
                    count=0
                    for i in range(1,4):
                        t=ws[col1+str(currRow+count)]
                        t.value=resValues.get(i)
                        cell_styling(t,"H")
                        count+=1
                else:
                    col1=get_column_letter(col)
                    count=0
                    for i in range(1,5):
                        t=ws[col1+str(currRow+count)]
                        t.value=resCost.get(i)
                        cell_styling(t,"H")
                        count+=1
            #Columns where data in inputed with a formula at the last row to display difference between the data inputed.
            elif(col>=4 and col<9):
                if(table!=2):
                    row=currRow-1
                    dayCol=get_column_letter(col) 
                    t=ws[dayCol+str(row)]
                    t.value=resValuesH.get(col)
                    cell_styling(t,"H")
                    t.fill=PatternFill(start_color='FFE49C', end_color='FFE49C', fill_type="solid")
                    meta_cell=dayCol+str(currRow)
                    res_cell=dayCol+str(currRow+1)
                    if(current_Supervisor!=0):
                        t=ws[meta_cell]
                        t.value="='"+sheetNames[0]+"'!"+prod_ref_cells[csr_Count]
                        cell_styling(t,"T")
                        csr_Count+=1
                        if(table==1):
                            difference_cell=ws[res_cell]
                            difference_cell.value="=AVERAGE("+dayCol+str(8)+":"+dayCol+str(org_row)+")"
                            cell_styling(difference_cell,"T")
                        difference_cell=ws[dayCol+str(currRow+2)] 
                        #------Value of cells equal the value of the first Supervisor, as the CSR is Department Wise------#
                        difference_cell.value="="+dayCol+str(currRow+1)+"-"+dayCol+str(currRow)
                        cell_styling(difference_cell,"T")
                    else:
                        if(table==1):
                            difference_cell=ws[res_cell]
                            difference_cell.value="=AVERAGE("+dayCol+str(8)+":"+dayCol+str(org_row)+")"
                            cell_styling(difference_cell,"T")
                        difference_cell=ws[dayCol+str(currRow+2)]
                        prod_ref_cells+=[meta_cell]
                        difference_cell.value="="+dayCol+str(currRow+1)+"-"+dayCol+str(currRow)
                        cell_styling(difference_cell,"T")
                else:
                    #Change cells to have formulas from Cost Table
                    # TODO LEFT AT CHOOSING WHAT TO DO DEPENDING ON ORDER OF SUPERVISOR
                    if(current_Supervisor!=0):
                        row=currRow-1
                        dayCol=get_column_letter(col) 
                        t=ws[dayCol+str(row)]
                        t.value=resValuesH.get(col)
                        cell_styling(t,"H")
                        t.fill=PatternFill(start_color='FFE49C', end_color='FFE49C', fill_type="solid")
                        meta_cell=org_row+3
                        r_cell=meta_cell+1
                        for i in range(0,4):
                            t=ws[dayCol+str(currRow+i)]
                            #----------------This is where the Nomina Cells are located-------#
                            if(i==0):
                                t.value="="+dayCol+str(currRow+1)+"/"+dayCol+str(meta_cell)
                            elif(i==2):
                                cost_ref_cells+=[dayCol+str(currRow+1)]
                                t.value="="+dayCol+str(currRow+1)+"/"+dayCol+str(r_cell)
                            elif(i==3):
                                t.value="="+dayCol+str(currRow)+"-"+dayCol+str(currRow+2)
                            cell_styling(t,"T")
                    else:
                        row=currRow-1
                        dayCol=get_column_letter(col) 
                        t=ws[dayCol+str(row)]
                        t.value=resValuesH.get(col)
                        cell_styling(t,"H")
                        t.fill=PatternFill(start_color='FFE49C', end_color='FFE49C', fill_type="solid")
                        meta_cell=org_row+3
                        r_cell=meta_cell+1
                        for i in range(0,4):
                            t=ws[dayCol+str(currRow+i)]
                            #----------------This is where the Nomina Cells are located-------#
                            if(i==0):
                                t.value="="+dayCol+str(currRow+1)+"/"+dayCol+str(meta_cell)
                            elif(i==2):
                                cost_ref_cells+=[dayCol+str(currRow+1)]
                                t.value="="+dayCol+str(currRow+1)+"/"+dayCol+str(r_cell)
                            elif(i==3):
                                t.value="="+dayCol+str(currRow)+"-"+dayCol+str(currRow+2)
                            cell_styling(t,"T")
                    
            #Works with the Weekly Eff. Column. This time working as a column to display cummulative work, eff, cost
            elif(col==10):
                if(table!=2):
                    row=currRow-1
                    dayCol=get_column_letter(col) 
                    t=ws[dayCol+str(row)]
                    t.value=resValues.get(len(resValues))
                    cell_styling(t,"H")
                    t.fill=PatternFill(start_color='F8B484', end_color='F8B484', fill_type="solid")
                    for i in range(1,4):
                        if(i!=3):
                            t=ws[dayCol+str(currRow)]
                            t.value="=SUM("+"D"+str(currRow)+":"+"H"+str(currRow)+")"
                            cell_styling(t,"H")
                            currRow+=1
                        else:
                            t=ws[dayCol+str(currRow)]
                            t.value="=$"+dayCol+"$"+str(currRow-1)+"-$"+dayCol+"$"+str(currRow-2)
                            cell_styling(t,"H")
                else:
                    row=currRow-1
                    dayCol=get_column_letter(col) 
                    t=ws[dayCol+str(row)]
                    t.value=resValues.get(len(resValues))
                    cell_styling(t,"H")
                    t.fill=PatternFill(start_color='F8B484', end_color='F8B484', fill_type="solid")
                    meta_cell=org_row+6
                    for i in range(0,4):
                        t=ws[dayCol+str(currRow+i)]
                        if(i==0):
                            t.value="=$"+dayCol+"$"+str(currRow+1)+"/$"+dayCol+"$"+str(meta_cell)
                        elif(i==2):
                            t.value="=$"+dayCol+"$"+str(currRow+1)+"/$"+dayCol+"$"+str(meta_cell+1)
                        elif(i==3):
                            t.value="=$"+dayCol+"$"+str(currRow)+"-$"+dayCol+"$"+str(currRow+2)
                        t.font=tableFont
                        t.alignment=alignCenter
                        t.border=thin_border
        currRow+=24
    print("\nCompleted Production Section Creation...")


                
def inputIDname(ws,iDs,employeeNames,effList,Departments,depCount,firstEntry):
    if(firstEntry<2):
        for col in range(1,5):
            count=0
            colLetter= get_column_letter(col) 
            weeklyCol=get_column_letter(10)
            for row in range(8,8+len(iDs)):
                t=ws[colLetter+str(row)]
                wt=ws[weeklyCol+str(row)]
                wt.value= '=AVERAGE(D'+str(row)+':H'+str(row)+')'
                wt.font=tableFont
                wt.alignment=alignCenter
                wt.border= thin_border
                if(col==1):
                    t.value=int(iDs[count])
                    cell_styling(t,"T")
                    count+=1
                elif(col==2):
                    t.value=Departments[depCount]
                    cell_styling(t,"T")
                elif(col==3):
                    t.value=employeeNames[count]
                    cell_styling(t,"T")
                    count+=1
                else:
                    t.value=float(effList[count])
                    cell_styling(t,"T")
                    count+=1

    else:
        col=get_column_letter(firstEntry+3)
        count=0
        weeklyCol=get_column_letter(10)
        for row in range(8,8+len(iDs)):
            wt=ws[weeklyCol+str(row)]
            wt.value= '=AVERAGE(D'+str(row)+':H'+str(row)+')'
            cell_styling(wt,"T")
            t=ws[col+str(row)]
            t.value=float(effList[count])
            cell_styling(t,"T")
            count+=1
    print("\nCompleting Employee Logging...")   

def fitTotext(firstEntry,ws,size):
    greatestWidth=0
    if(firstEntry<2):
        for col in range(1,11):
            if(col==3):
                for row in range(8,size):
                    cell=ws[get_column_letter(col)+str(row)]
                    if(greatestWidth<len(cell.value)):
                        greatestWidth=len(cell.value)
                greatestWidth=greatestWidth*2.5
                ws.column_dimensions[get_column_letter(col)].width=greatestWidth
            elif(col==9):
                greatestWidth=1.25
                ws.column_dimensions[get_column_letter(col)].width=greatestWidth
            else:
                greatestWidth=21.5
                ws.column_dimensions[get_column_letter(col)].width=greatestWidth
    else:
        greatestWidth=21.5
        ws.column_dimensions[get_column_letter(firstEntry+3)].width=greatestWidth

def cell_styling(cell, style):
        if(style=="H"): #Header Font
            cell.border=thin_border
            cell.font=headFont
            cell.alignment=alignCenter
        elif(style=="T"): #Table Font
            cell.border=thin_border
            cell.font=tableFont
            cell.alignment=alignCenter

if __name__ == "__main__":
    main()