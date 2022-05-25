import Main
import unittest
from openpyxl import Workbook, workbook, load_workbook

class Test(unittest.TestCase):
    def test_csr_creation(self):
        Report=load_workbook("results/Daily Eff. Report AutoTest copy.xlsx")
        ws=Report["NAHILA 22-MAY"]
        Original=load_workbook("results/OPR AGSU MAY 20.xlsx")
        Main.create_prod_rep(ws,23,0,"NAHILA 22-MAY")

def main():
    Test.test_csr_creation()
