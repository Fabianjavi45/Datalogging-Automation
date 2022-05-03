#imports
from calendar import month
from gc import collect
from pickle import FALSE, NONE
from openpyxl import Workbook, workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Fill, NamedStyle
from openpyxl.styles.numbers import FORMAT_DATE_XLSX15, FORMAT_DATE_DDMMYY
from openpyxl.styles.colors import Color
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.utils import get_column_letter
from  collections import defaultdict
import Date

#Global Variables and Resources
res={
    "Title":{
        1:"PRIFB",
        2:"OPERATOR EFFICIENCY GREEN SERVICE ",
        3:"WEEK OF: ",
        4:"SUPERVISOR: "
    },
    "Heading":{
        1:"Employee",
        2:"Department",
        3:"Employee Name",
        4:"LUNES",
        5:"MARTES",
        6:"MIERCOLES",
        7:"JUEVES",
        8:"VIERNES",
        10:"WEEKLY EFF."
    },
    "Months":{
        1:"JANUARY",
        2:"FEBRUARY",
        3:"MARCH",
        4:"APRIL",
        5:"MAY",
        6:"JUNE",
        7:"JULY",
        8:"AUGUST",
        9:"SEPTEMBER",
        10:"OCOTOBER",
        11:"NOVEMBER",
        12:"DECEMBER"
    },
    "TabColors":{
        1:"BC9CFF",
        2:"84F8B4",
        3:"84C8F8",
        4:"F88486"
    },
    "Prod_Eff":{
        1:"META",
        2:"RESULTADO",
        3:"DIFERENCIA",
        4:"ACUMULATIVO"
    },
    "Cost":{
        1:"META",
        2:"NOMINA (TPM)",
        3:"COSTO POR UNIDAD",
        4:"DIFERENCIA",
        5:"ACUMULATIVO"
    },
    "Prod_Cost_Title":{
        0:"PRODUCCION",
        1:"EFICIENCIA",
        2:"COSTO"
    }
}
headFont= Font(name="Verdana",sz= 24, bold=True)
alignCenter= Alignment(horizontal="center")
tableFont= Font(name="Verdana",sz=24) 
thin_border= Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
TabColor=2

class autoxl():
    def _init_(this,original,result):
        this.original=original
        this.result=result
    #Gets the ID's of every employee in the original report
    def getIds(this,row):
        ows=this.original.active                                 
        col="A"
        t=ows[col+str(row)]
        idList=[]
        while(t.value[0]!="T"):
            id=t.value.split(' ',1)
            idList=idList + [int(id[0])]
            row+=6
            t=ows[col+str(row)]
            while(t.value is None):
                row+=1
                t=ows[col+str(row)]
        return idList
    
    def getchildId(this,ws):
        resultEff=[]
        col="A"
        row=8
        t=ws[col+str(row)]
        while(t.value != None):
            resultEff=resultEff+[t.value]
            row+=1
            t=ws[col+str(row)]
        return resultEff
    #Gets the names of all the employees in the original report
    def setNewEntries(this,ws,employeeNames,oldCount,iDs):
        rowtoInsert=8+oldCount
        ws.insert_rows(rowtoInsert)
        t=ws["A"+str(rowtoInsert)]
        t.value=int(iDs[oldCount])
        t.font=tableFont
        t.alignment=alignCenter
        t.border= thin_border
        t=ws["B"+str(rowtoInsert)]
        if(ws["B"+str(rowtoInsert+1)].value==None):
            t.value=ws["B"+str(rowtoInsert-1)].value
        else:
            t.value=ws["B"+str(rowtoInsert+1)].value
        t.font=tableFont
        t.alignment=alignCenter
        t.border= thin_border
        t=ws["C"+str(rowtoInsert)]
        t.value=employeeNames[oldCount]
        t.font=headFont
        t.alignment=alignCenter
        t.border= thin_border
        return

    def getNames(this,row):
        ows=this.original.active                               
        col="A"
        t=ows[col+str(row)]
        nameList=[]
        while(t.value[0]!="T"):
            nameStart=t.value.split(')')
            finalName=nameStart[1].split('Supervisor')
            nameList= nameList + [finalName[0]]
            row+=6
            t=ows[col+str(row)]
            while(t.value is None):
                row+=1
                t=ows[col+str(row)]
        return nameList

    def getEff(this,row):
        ows=this.original.active
        col="L"
        t=ows[col+str(row)]
        effList=[]
        while(t.value[0]!="T"):
            splitEff=t.value.split("\xa0\xa0")
            splitEff=splitEff[1].split('%')
            effList= effList + [float(splitEff[0])]
            row+=6
            t=ows[col+str(row)]
            last=ows["A"+str(row-1)]
            offStandard=ows["D"+str(row)].value
            while(t.value is None or t.value==ows["A"+str(row)] or offStandard == None or isinstance(offStandard,str)):
                last=ows["A"+str(row)]
                row+=1
                offStandard=ows["D"+str(row)].value
                t=ows[col+str(row)]
            if(last.value != None and last.value[0]=="T"):
                t=last
        return effList

    def getDepart(this):
        ows=this.original.active
        t=ows["A"+str(4)]
        firstsplit=t.value.split('Department:')
        internalDepart=firstsplit[1].split(',')
        return internalDepart

    def getSupers(this,numOfdep):
        ows=this.original.active
        row=9                                 #List always starts at this cell number
        col="A"
        t=ows[col+str(row)]
        supList=[]
        for i in range(0,numOfdep):
            sup=t.value.split('Supervisor:')
            sup=sup[1].split(" ")
            #sup1=sup[1]+" "+sup[2]
            supList=supList + [sup[1]]
            if(i==numOfdep-1):
                break
            while(t.value[0]!="T"):
                row+=6
                t=ows[col+str(row)]
                while(t.value is None):
                    row+=1
                    t=ows[col+str(row)]           
            row+=7
            t=ows[col+str(row)]
        return supList
    
    def getdeptDiv(this,numOfdep):
        Divisionrows=[9]
        ows=this.original.active
        row=9                                 #List always starts at this cell number
        col="A"
        t=ows[col+str(row)]
        for i in range(0,numOfdep):
            if(i==numOfdep-1):
                break
            while(t.value[0]!="T"):
                row+=6
                t=ows[col+str(row)]
                while(t.value is None):
                    row+=1
                    t=ows[col+str(row)]           
            row+=7
            Divisionrows=Divisionrows+[row]
            t=ows[col+str(row)]
        return Divisionrows

    def getDay(this):
        ows=this.original.active
        row=3
        col="A"
        t=ows[col+str(row)]
        firstSplit=t.value.split("{")
        date=firstSplit[1].split("}")
        date=date[0]
        return date

    def input_dead_day(this,ws,iDs,riDs,effList,firstEntry):
        col=get_column_letter(firstEntry+3)
        for i in range(0, len(iDs)):
            if(riDs.index(iDs[i])!=ValueError):
                row=riDs.index(iDs[i])+8
                t=ws[col+str(row)]
                t.value=float(effList[i])
            t.font=tableFont
            t.alignment=alignCenter
            t.border= thin_border   





