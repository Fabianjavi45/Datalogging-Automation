#imports
from calendar import month
from gc import collect
from pickle import FALSE, NONE
from openpyxl import Workbook, workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Fill, NamedStyle
from openpyxl.styles.numbers import FORMAT_DATE_XLSX15, FORMAT_DATE_DDMMYY
from openpyxl.styles.colors import Color
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.utils import get_column_letter
from  collections import defaultdict

#Global Variables and Resources
res={
    "Title":{
        1:"PRIFB",
        2:"OPERATOR EFFICIENCY GREEN SERVICE ",
        3:"WEEK OF: ",
        4:"SUPERVISOR: "
    },
    "Heading":{
        1:"Employee",
        2:"Department",
        3:"Employee Name",
        4:"LUNES",
        5:"MARTES",
        6:"MIERCOLES",
        7:"JUEVES",
        8:"VIERNES",
        10:"WEEKLY EFF."
    },
    "Months":{
        1:"JANUARY",
        2:"FEBRUARY",
        3:"MARCH",
        4:"APRIL",
        5:"MAY",
        6:"JUNE",
        7:"JULY",
        8:"AUGUST",
        9:"SEPTEMBER",
        10:"OCOTOBER",
        11:"NOVEMBER",
        12:"DECEMBER"
    },
    "TabColors":{
        1:"BC9CFF",
        2:"84F8B4",
        3:"84C8F8",
        4:"F88486"
    },
    "Prod_Eff":{
        1:"META",
        2:"RESULTADO",
        3:"DIFERENCIA",
        4:"ACUMULATIVO"
    },
    "Cost":{
        1:"META",
        2:"NOMINA (TPM)",
        3:"COSTO POR UNIDAD",
        4:"DIFERENCIA",
        5:"ACUMULATIVO"
    },
    "Prod_Cost_Title":{
        0:"PRODUCCION",
        1:"EFICIENCIA",
        2:"COSTO"
    }
}
headFont= Font(name="Verdana",sz= 24, bold=True)
alignCenter= Alignment(horizontal="center")
tableFont= Font(name="Verdana",sz=24) 
thin_border= Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
TabColor=2

class autoxl():
    def _init_(this,original,result):
        this.original=original
        this.result=result
    #Gets the ID's of every employee in the original report
    def getIds(this,row):
        ows=this.original.active                                 
        col="A"
        t=ows[col+str(row)]
        idList=[]
        while(t.value[0]!="T"):
            id=t.value.split(' ',1)
            idList=idList + [int(id[0])]
            row+=6
            t=ows[col+str(row)]
            while(t.value is None):
                row+=1
                t=ows[col+str(row)]
        return idList
    
    def getchildId(this,ws):
        resultEff=[]
        col="A"
        row=8
        t=ws[col+str(row)]
        while(t.value != None):
            resultEff=resultEff+[t.value]
            row+=1
            t=ws[col+str(row)]
        return resultEff
    #Gets the names of all the employees in the original report
    def setNewEntries(this,ws,employeeNames,oldCount,iDs):
        rowtoInsert=8+oldCount
        ws.insert_rows(rowtoInsert)
        t=ws["A"+str(rowtoInsert)]
        t.value=int(iDs[oldCount])
        t.font=tableFont
        t.alignment=alignCenter
        t.border= thin_border
        t=ws["B"+str(rowtoInsert)]
        if(ws["B"+str(rowtoInsert+1)].value==None):
            t.value=ws["B"+str(rowtoInsert-1)].value
        else:
            t.value=ws["B"+str(rowtoInsert+1)].value
        t.font=tableFont
        t.alignment=alignCenter
        t.border= thin_border
        t=ws["C"+str(rowtoInsert)]
        t.value=employeeNames[oldCount]
        t.font=headFont
        t.alignment=alignCenter
        t.border= thin_border
        return

    def getNames(this,row):
        ows=this.original.active                               
        col="A"
        t=ows[col+str(row)]
        nameList=[]
        while(t.value[0]!="T"):
            nameStart=t.value.split(')')
            finalName=nameStart[1].split('Supervisor')
            nameList= nameList + [finalName[0]]
            row+=6
            t=ows[col+str(row)]
            while(t.value is None):
                row+=1
                t=ows[col+str(row)]
        return nameList

    def getEff(this,row):
        ows=this.original.active
        col="L"
        t=ows[col+str(row)]
        effList=[]
        while(t.value[0]!="T"):
            splitEff=t.value.split("\xa0\xa0")
            splitEff=splitEff[1].split('%')
            effList= effList + [float(splitEff[0])]
            row+=6
            t=ows[col+str(row)]
            last=ows["A"+str(row-1)]
            offStandard=ows["D"+str(row)].value
            while(t.value is None or t.value==ows["A"+str(row)] or offStandard == None or isinstance(offStandard,str)):
                last=ows["A"+str(row)]
                row+=1
                offStandard=ows["D"+str(row)].value
                t=ows[col+str(row)]
            if(last.value != None and last.value[0]=="T"):
                t=last
        return effList

    def getDepart(this):
        ows=this.original.active
        t=ows["A"+str(4)]
        firstsplit=t.value.split('Department:')
        internalDepart=firstsplit[1].split(',')
        return internalDepart

    def getSupers(this,numOfdep):
        ows=this.original.active
        row=9                                 #List always starts at this cell number
        col="A"
        t=ows[col+str(row)]
        supList=[]
        for i in range(0,numOfdep):
            sup=t.value.split('Supervisor:')
            sup=sup[1].split(" ")
            #sup1=sup[1]+" "+sup[2]
            supList=supList + [sup[1]]
            if(i==numOfdep-1):
                break
            while(t.value[0]!="T"):
                row+=6
                t=ows[col+str(row)]
                while(t.value is None):
                    row+=1
                    t=ows[col+str(row)]           
            row+=7
            t=ows[col+str(row)]
        return supList
    
    def getdeptDiv(this,numOfdep):
        Divisionrows=[9]
        ows=this.original.active
        row=9                                 #List always starts at this cell number
        col="A"
        t=ows[col+str(row)]
        for i in range(0,numOfdep):
            if(i==numOfdep-1):
                break
            while(t.value[0]!="T"):
                row+=6
                t=ows[col+str(row)]
                while(t.value is None):
                    row+=1
                    t=ows[col+str(row)]           
            row+=7
            Divisionrows=Divisionrows+[row]
            t=ows[col+str(row)]
        return Divisionrows

    def getDay(this):
        ows=this.original.active
        row=3
        col="A"
        t=ows[col+str(row)]
        firstSplit=t.value.split("{")
        date=firstSplit[1].split("}")
        date=date[0]
        return date

    def input_dead_day(this,ws,iDs,riDs,effList,firstEntry):
        col=get_column_letter(firstEntry+3)
        for i in range(0, len(iDs)):
            if(riDs.index(iDs[i])!=ValueError):
                row=riDs.index(iDs[i])+8
                t=ws[col+str(row)]
                t.value=float(effList[i])
            t.font=tableFont
            t.alignment=alignCenter
            t.border= thin_border   

class Date():

    def __init__(this,day,month,year):
        this.day=day
        this.month=month
        this.year=year

    #Gets the week 
    def getWeek(this):
        if(this.month==1 or this.month==3 or this.month==5 or this.month==7 
        or this.month==8 or this.month==10 or this.month==12):
            if(this.day>25):
                newDay=6-(31-this.day)
                newMonth=this.month+1
                newWeek= Date(newDay,newMonth,this.year)
                return newWeek
        if(this.month==4 or this.month==6 or this.month==9 or this.month==9 or this.month==11):
            if(this.day>24):
                newDay=6-(30-this.day)
                newMonth=this.month+1
                newWeek= Date(newDay,newMonth,this.year)
                return newWeek
            
        if(this.month==2):
            if(this.day>22):
                newDay=6-(28-this.day)
                newMonth=this.month+1
                newWeek= Date(newDay,newMonth,this.year)
                return newWeek
        newDay=this.day+6
        return Date(newDay,this.month,this.year)

    def getendWeek(this,weekDay):
        if(weekDay=="M"):
            weekDay=2
        elif(weekDay=="W"):
            weekDay=3
        elif(weekDay=="J"):
            weekDay=4
        elif(weekDay=="V"):
            weekDay=5
        
        if(this.month==1 or this.month==3 or this.month==5 or this.month==7 
        or this.month==8 or this.month==10 or this.month==12):
            if(this.day>25):
                newDay=(7-weekDay)+this.day
                if(newDay>31):
                    newDay=(7-weekDay)-(31-this.day)
                    newMonth=this.month+1
                    newWeek= Date(newDay,newMonth,this.year)
                    return newWeek
                newWeek= Date(newDay,this.month,this.year)
                return newWeek
            else:
                newDay=(7-weekDay)+this.day
                endWeek= Date(newDay,this.month,this.year)
                return endWeek
        if(this.month==4 or this.month==6 or this.month==9 or this.month==9 or this.month==11):
            if(this.day>24):
                newDay=(7-weekDay)+this.day
                if(newDay>30):
                    newDay=(7-weekDay)-(30-this.day)
                    newMonth=this.month+1
                    newWeek= Date(newDay,newMonth,this.year)
                    return newWeek
                newWeek= Date(newDay,this.month,this.year)
                return newWeek
            else:
                newDay=(7-weekDay)+this.day
                endWeek= Date(newDay,this.month,this.year)
                return endWeek    
        if(this.month==2):
            if(this.day>22):
                newDay=(7-weekDay)+this.day
                if(newDay>28):
                    newDay=(7-weekDay)-(28-this.day)
                    newMonth=this.month+1
                    newWeek= Date(newDay,newMonth,this.year)
                    return newWeek
                newWeek= Date(newDay,this.month,this.year)
                return newWeek
            else:
                newDay=(7-weekDay)+this.day
                endWeek= Date(newDay,this.month,this.year)
                return endWeek
    def getdayNumber(this,weekDay):
        if(weekDay=="M"):
            weekDay=2
        if(weekDay=="W"):
            weekDay=3
        if(weekDay=="J"):
            weekDay=4
        if(weekDay=="V"):
            weekDay=5
        return weekDay

def startUp():
    print('\n\n********This program is to simplify the Employee Reports of PRIFB & HARDWICK and automatically make an Efficiency Report********\n')
    print('First please input the file names to work with.\n')  

def createTitle(ws,weekDate,endweekDate,Departments,supName):
    resValues=res["Title"]
    tMonths=res["Months"]
    print("Starting Title Creation...")
    for row in range(1,5):
        char1=get_column_letter(1)
        char2=get_column_letter(10)
        #print(char1,char2)
        ws.merge_cells(char1+str(row)+":"+char2+str(row))
        t=ws[char1+str(row)]
        if(row==3):
            t.value=resValues.get(row)+str(weekDate.day)+"-"+tMonths.get(weekDate.month)+"-"+str(weekDate.year)+" TO "+str(endweekDate.day)+"-"+tMonths.get(endweekDate.month)+"-"+str(endweekDate.year)
        elif(row==2):
            t.value=resValues.get(row)+"("+Departments[0]+")" #Departments[0] should be changed for when creating multiple sheets
        elif(row==4):
            t.value=resValues.get(row)+supName
        else:
            t.value=resValues.get(row)
        t.font=headFont
        t.alignment=alignCenter
    ws.pageSetUpPr = PageSetupProperties(fitToPage=True)
    print("Completed Title Creation...")

def createHeader(ws,weekDate):
    print("Starting Heading Creation...")
    #weekDate=Date(28,2,2022)
    resValues=res["Heading"]
    dayCount=weekDate.day
    for col in range(1,11):
        row=5
        if(col<4 or col>9):
            scol=get_column_letter(col)
            ws.merge_cells(scol+str(row)+":"+scol+str(row+2))
            t=ws[scol+str(row)]
            t.value=resValues.get(col)
            if(col>9):
                t.fill=PatternFill(start_color='F8B484', end_color='F8B484', fill_type="solid")
            else:
                t.fill=PatternFill(start_color='FFE49C', end_color='FFE49C', fill_type="solid")
            t.border=thin_border
        elif(col<9):
            for row in range(5,8):
                if(row==5):
                    #write day
                    dayCol=get_column_letter(col) 
                    t=ws[dayCol+str(row)]
                    t.value=resValues.get(col)
                elif(row==6):
                    #write date
                    t=ws[dayCol+str(row)]
                    if(col>4):
                        t.style="dateformat"
                        t.value="="+get_column_letter(col-1)+str(row)+"+1"
                    else:
                        t.style="dateformat"
                        t.value=str(weekDate.month)+"/"+str(dayCount)+"/"+str(weekDate.year)
                else:
                    t=ws[dayCol+str(row)]
                    t.value="Eff."
                t.font=headFont
                t.alignment=alignCenter
                t.fill=PatternFill(start_color='FFE49C', end_color='FFE49C', fill_type="solid")
                t.border= thin_border
        t.font=headFont
        t.alignment=alignCenter
    print("Completed Heading Creation...")  
#Creates the following 3 tables that display Production, Line Efficiency and Cost per Unit

def create_prod_rep(ws, currRow):
    print("Starting Production Section Creation...")
    resValues=res["Prod_Eff"]
    resValuesH=res["Heading"]
    resTable=res["Prod_Cost_Title"]
    resCost=res["Cost"]
    org_row=currRow+7
    currRow+=13
    meta_cell=" "

    for table in range(0,3):
        for col in range(1,11):
            if(col==1):
                col1=get_column_letter(col)
                col2=get_column_letter(col+1)
                ws.merge_cells(col1+str(currRow)+":"+col2+str(currRow+2))
                t=ws[col1+str(currRow)]
                t.value=resTable.get(table)
                t.font=headFont
                t.alignment=alignCenter
                t.border= thin_border
            elif(col==3):
                if(table!=2):
                    col1=get_column_letter(col)
                    count=0
                    for i in range(1,4):
                        t=ws[col1+str(currRow+count)]
                        t.value=resValues.get(i)
                        count+=1
                        t.font=headFont
                        t.alignment=alignCenter
                        t.border= thin_border
                else:
                    col1=get_column_letter(col)
                    count=0
                    for i in range(1,5):
                        t=ws[col1+str(currRow+count)]
                        t.value=resCost.get(i)
                        count+=1
                        t.font=headFont
                        t.alignment=alignCenter
                        t.border= thin_border
            #Columns where data in inputed with a formula at the last row to display difference between the data inputed.
            elif(col>=4 and col<9):
                if(table!=2):
                    row=currRow-1
                    dayCol=get_column_letter(col) 
                    t=ws[dayCol+str(row)]
                    t.value=resValuesH.get(col)
                    t.font=headFont
                    t.alignment=alignCenter
                    t.border=thin_border
                    t.fill=PatternFill(start_color='FFE49C', end_color='FFE49C', fill_type="solid")
                    meta_cell=dayCol+str(currRow)
                    res_cell=dayCol+str(currRow+1)
                    if(table==1):
                        r_cell=ws[res_cell]
                        r_cell.value="=AVERAGE("+dayCol+str(8)+":"+dayCol+str(org_row)+")"
                        r_cell.border=thin_border
                        r_cell.font=tableFont
                        r_cell.alignment=alignCenter
                    curr_cell=ws[dayCol+str(currRow+2)]
                    curr_cell.value="="+dayCol+str(currRow+1)+"-"+dayCol+str(currRow)
                    curr_cell.border=thin_border
                    curr_cell.font=tableFont
                    curr_cell.alignment=alignCenter
                else:
                    row=currRow-1
                    dayCol=get_column_letter(col) 
                    t=ws[dayCol+str(row)]
                    t.value=resValuesH.get(col)
                    t.font=headFont
                    t.alignment=alignCenter
                    t.border=thin_border
                    t.fill=PatternFill(start_color='FFE49C', end_color='FFE49C', fill_type="solid")
                    meta_cell=org_row+6
                    for i in range(0,4):
                        t=ws[dayCol+str(currRow+i)]
                        if(i==0):
                            t.value="="+dayCol+str(currRow+1)+"/"+dayCol+str(meta_cell)
                        elif(i==2):
                           t.value="="+dayCol+str(currRow+1)+"/"+dayCol+str(meta_cell+1)
                        elif(i==3):
                            t.value="="+dayCol+str(currRow)+"-"+dayCol+str(currRow+2)
                        t.font=tableFont
                        t.alignment=alignCenter
                        t.border=thin_border
                    #TODO Change cells to have formulas from Cost Table
            #Works with the Weekly Eff. Column. This time working as a column to display cummulative work, eff, cost
            elif(col==10):
                if(table!=2):
                    row=currRow-1
                    dayCol=get_column_letter(col) 
                    t=ws[dayCol+str(row)]
                    t.value=resValues.get(len(resValues))
                    t.font=headFont
                    t.alignment=alignCenter
                    t.border=thin_border
                    t.fill=PatternFill(start_color='F8B484', end_color='F8B484', fill_type="solid")
                    for i in range(1,4):
                        if(i!=3):
                            t=ws[dayCol+str(currRow)]
                            t.value="=SUM("+"D"+str(currRow)+":"+"H"+str(currRow)+")"
                            currRow+=1
                            t.font=headFont
                            t.alignment=alignCenter
                            t.border= thin_border
                        else:
                            t=ws[dayCol+str(currRow)]
                            t.value="=$"+dayCol+"$"+str(currRow-1)+"-$"+dayCol+"$"+str(currRow-2)
                            t.font=headFont
                            t.alignment=alignCenter
                            t.border= thin_border
                else:
                    row=currRow-1
                    dayCol=get_column_letter(col) 
                    t=ws[dayCol+str(row)]
                    t.value=resValues.get(len(resValues))
                    t.font=headFont
                    t.alignment=alignCenter
                    t.border=thin_border
                    t.fill=PatternFill(start_color='F8B484', end_color='F8B484', fill_type="solid")
                    meta_cell=org_row+6
                    for i in range(0,4):
                        t=ws[dayCol+str(currRow+i)]
                        if(i==0):
                            t.value="=$"+dayCol+"$"+str(currRow+1)+"/$"+dayCol+"$"+str(meta_cell)
                        elif(i==2):
                           t.value="=$"+dayCol+"$"+str(currRow+1)+"/$"+dayCol+"$"+str(meta_cell+1)
                        elif(i==3):
                            t.value="=$"+dayCol+"$"+str(currRow)+"-$"+dayCol+"$"+str(currRow+2)
                        t.font=tableFont
                        t.alignment=alignCenter
                        t.border=thin_border
        currRow+=24
                
def inputIDname(ws,iDs,employeeNames,effList,Departments,depCount,firstEntry):
    if(firstEntry<2):
        for col in range(1,5):
            count=0
            colLetter= get_column_letter(col) 
            weeklyCol=get_column_letter(10)
            for row in range(8,8+len(iDs)):
                t=ws[colLetter+str(row)]
                wt=ws[weeklyCol+str(row)]
                wt.value= '=AVERAGE(D'+str(row)+':H'+str(row)+')'
                wt.font=tableFont
                wt.alignment=alignCenter
                wt.border= thin_border
                if(col==1):
                    t.value=int(iDs[count])
                    t.font=tableFont
                    t.alignment=alignCenter
                    t.border= thin_border
                    count+=1
                elif(col==2):
                    t.value=Departments[depCount]
                    t.font=tableFont
                    t.alignment=alignCenter
                    t.border= thin_border
                elif(col==3):
                    t.value=employeeNames[count]
                    t.font=headFont
                    t.alignment=alignCenter
                    t.border= thin_border
                    count+=1
                else:
                    t.value=float(effList[count])
                    t.font=tableFont
                    t.alignment=alignCenter
                    t.border= thin_border
                    count+=1

    else:
        col=get_column_letter(firstEntry+3)
        count=0
        weeklyCol=get_column_letter(10)
        for row in range(8,8+len(iDs)):
            wt=ws[weeklyCol+str(row)]
            wt.value= '=AVERAGE(D'+str(row)+':H'+str(row)+')'
            wt.font=tableFont
            wt.alignment=alignCenter
            wt.border= thin_border
            t=ws[col+str(row)]
            t.value=float(effList[count])
            t.font=tableFont
            t.alignment=alignCenter
            t.border= thin_border
            count+=1
    print("\nCompleting Employee Logging...")   

def fitTotext(firstEntry,ws,size):
    greatestWidth=0
    if(firstEntry<2):
        for col in range(1,11):
            if(col==3):
                for row in range(8,size):
                    cell=ws[get_column_letter(col)+str(row)]
                    if(greatestWidth<len(cell.value)):
                        greatestWidth=len(cell.value)
                greatestWidth=greatestWidth*2.5
                ws.column_dimensions[get_column_letter(col)].width=greatestWidth
            elif(col==9):
                greatestWidth=1.25
                ws.column_dimensions[get_column_letter(col)].width=greatestWidth
            else:
                greatestWidth=21.5
                ws.column_dimensions[get_column_letter(col)].width=greatestWidth
    else:
        greatestWidth=21.5
        ws.column_dimensions[get_column_letter(firstEntry+3)].width=greatestWidth

def main(orgReport, resReport, Day): #orgReport, resReport, Day
    #startUp()
    #---------------Intial input of the two workbooks to work with-----------------#
    orgWB= orgReport #"original/OPR MCTR APR 11.xlsx"
    resWB= resReport#"results/Daily Eff. Report AutoTest.xlsx" #resReport
    #---------------Initializing Global Variables---------------------------#
    p1=autoxl()
    p1.original=load_workbook(orgWB)
    p1.result=load_workbook(resWB)
    Departments=p1.getDepart()
    Supervisors=p1.getSupers(len(Departments))
    Divisionrows=p1.getdeptDiv(len(Departments))
    #-------------------------Mode Select---------------------------------#
    print("\n\nPROGRAM START...\n\n")
    weekDay=Day #"L" #Day #input("What day of the week is it? Please input day as: M=Martes, W==Miercoles, J=Jueves or V=Viernes: ")
    #Modeselect='n'#input('\First entry of the week? Answer: y or n')

    if(weekDay=="L"):
        firstEntry=1
        date=p1.getDay()
        date=date.split("/")
        weekDate=Date(int(date[1]),int(date[0]),2000+int(date[2]))
        endweekDate=weekDate.getWeek()

        #----------------Style for first entry of week-----------------------#
        #dateStyle=NamedStyle(name="dateformat",number_format=FORMAT_DATE_XLSX15)
        #p1.result.add_named_style(dateStyle)
        #--------------------------------------------------------------------#
        
        #Create Sheet and specifications
        sheetDate=" "+str(endweekDate.day)+"-"+res["Months"].get(endweekDate.month)
        sheetNames=[]
        #TODO Input the correct emplyees on the correct worksheets

        for i in range(0,len(Supervisors)):
            firstName=Supervisors[i].upper()
            sheetNames=sheetNames+[firstName+sheetDate]
            p1.result.create_sheet(sheetNames[i])
            ws=p1.result[sheetNames[i]]
            wsProp=ws.sheet_properties
            wsProp.tabColor=res["TabColors"].get(TabColor)

            #Sheet Title Creation
            createTitle(ws,weekDate,endweekDate,Departments,firstName)
            #Header Creation
            createHeader(ws,weekDate)

            print("\nStarting Employee Logging...")    
            iDs=p1.getIds(Divisionrows[i])
            employeeNames=p1.getNames(Divisionrows[i])
            effList=p1.getEff(Divisionrows[i]+1)

            #---------------Input new day entries---------------------------#
            inputIDname(ws,iDs,employeeNames,effList,Departments,i,firstEntry)
            #---------------------------------------------------------------# 
            #-------------------AutoFit Columns-----------------------------#
            fitTotext(firstEntry,ws,len(effList))  

            #Create Production, Overall Efficiency and Work Line Cost Table Reports
            #currRow=len(effList)
            #create_prod_rep(ws, currRow)


        #-----end for-------------------------------------------------------#  
        



        p1.result.save(resWB)
        print("\n\nPROGRAM END...")
    else:
        #Ask what day of the week is L=Lunes, M=Martes, W=Miercoles, J=Jueves, V=Viernes
        date=p1.getDay()
        date=date.split("/")
        weekDate=Date(int(date[1]),int(date[0]),2000+int(date[2]))
        dayNum= weekDate.getdayNumber(weekDay)
        endweekDate=weekDate.getendWeek(weekDay)
        #-----------GET THE NAMES OF THE SHEETS TO EDIT----------------------------#
        sheetDate=" "+str(endweekDate.day)+"-"+res["Months"].get(endweekDate.month)
        sheetNames=[]
        for i in range(0,len(Supervisors)):
            two_diff=False
            firstName=Supervisors[i].upper()
            sheetNames=sheetNames+[firstName+sheetDate] #Gets the name of the sheet based on the supervisors in the original report
            ws=p1.result[sheetNames[i]]
            print("\nStarting Employee Logging...")    
            iDs=p1.getIds(Divisionrows[i])
            riDs=p1.getchildId(ws)
            employeeNames=p1.getNames(Divisionrows[i])
            effList=p1.getEff(Divisionrows[i]+1)
            DEBUG=0
            #----------ID comparison to see if new entries were made----------#
            iCount=0
            if(len(iDs)*2<len(riDs)):
                #---------Implementation of Dead Day, in case of very few employee Antendance--------------#
                p1.input_dead_day(ws,iDs,riDs,effList,dayNum)
            else:
                for (newElement,oldElement) in zip(iDs,riDs):
                    if(two_diff):
                        two_diff=False
                        continue
                        
                    #Check if a new entry has to be added by comparing the positions of the current 
                    # elements of each list
                    if(oldElement>newElement):
                        p1.setNewEntries(ws,employeeNames,iCount,iDs)
                        DEBUG+=1
                        riDs.insert(iCount,newElement)
                        iCount+=1
                    #Checks if a new entry is already in the Eff. Report, If so then this new entry 
                    # is added to the result Report
                    elif(oldElement<newElement):
                        try:
                            iDs.index(oldElement)
                        except:
                            effList.insert(iCount,0)
                            iDs.insert(iCount,oldElement)
                            employeeNames.insert(iCount," ")
                        try:
                            riDs.index(newElement)
                        except:
                            iCount+=1
                            p1.setNewEntries(ws,employeeNames,iCount,iDs)
                            DEBUG+=1
                            riDs.insert(iCount,newElement)
                            continue
                        #diff=abs(riDs.index(newElement)-iDs.index(newElement))
                        iCount+=1
                    else:
                        iCount+=1
                #---------------Check for Data Inconsistencies---------------------------#
                if(len(iDs)>len(riDs)):
                    diff=len(iDs)-len(riDs)
                    for i in range(0,diff):
                        p1.setNewEntries(ws,employeeNames,iCount,iDs)
                        DEBUG+=1
                        riDs.insert(iCount,iDs[iCount])
                        iCount+=1
                if(len(iDs)!=len(riDs)):
                    #Raise ValueError so that the workbook isn't saved with corrupted data
                    #raise ValueError('There were inconsistencies in the data to input with the data already in file')
                    print
                #---------------Input new day entries---------------------------#    
                inputIDname(ws,iDs,employeeNames,effList,Departments,i,dayNum)
                #---------------------------------------------------------------#
            #-------------------AutoFit Columns-----------------------------#
            fitTotext(dayNum, ws,len(effList))
        #-----end for-------------------------------------------------------#
        p1.result.save(resWB)
        if(weekDay=='V'):
           currRow=len(riDs)
           create_prod_rep(ws, currRow) 
        print("\n\nPROGRAM END...")    

if __name__ == "__main__":
    main()

